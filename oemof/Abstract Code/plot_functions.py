import matplotlib.pyplot as plt
import matplotlib.gridspec as gridspec
import matplotlib.ticker as tkr

from bokeh.plotting import figure, show
from bokeh.layouts import layout
from bokeh.models import DatetimeTickFormatter
from bokeh.palettes import *
from bokeh.embed import file_html
from bokeh.resources import CDN

from openpyxl import load_workbook

import itertools
import pandas as pd

# This file defines different functions for the plotting of the results of the optimization.
# The plots are made at the end of this file, introducing a .xls file previously created during the optimization.

def monthly_balance(data, bus, new_legends):
    """
    Function for the definition of the monthly summary of a bus
    :param data: dict type, results from the optimization applied to one bus
    :param bus: str type, bus from which the summary is required
    :param new_legends: dict type, new legends to plot on the graph
    :return:
    """
    l = []
    for i in bus:
        l.append(i)
    l.reverse()
    l = l[0:11]  # in the case of a maximum of 9 buildings
    l.reverse()
    building = ''.join([str(item) for item in l])

    data_month = data.resample('1m').sum()
    monthShortNames = ['Jan', 'Fev', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']

    while len(data_month) != len(monthShortNames) :
        data.drop(data.index[-1], inplace=True)
        data_month = data.resample('1m').sum()
    neg_flow = []
    pos_flow = []
    for i in data.columns:
        a = [i.strip("()").split(", ")]
        if bus in a[0][0]:
            neg_flow.append(i)
        else:
            pos_flow.append(i)

    plt.figure()
    mark = []
    for i in neg_flow:
        plt.bar(monthShortNames, -data_month[i], label=new_legends[i.replace(building, "")], bottom=sum(mark))
        mark.append(-data_month[i])
    mark = []
    for i in pos_flow:
        plt.bar(monthShortNames, data_month[i], label=new_legends[i.replace(building, "")], bottom=sum(mark))
        mark.append(data_month[i])

    ax = plt.gca()
    # Shrink current axis by 20%
    box = ax.get_position()
    ax.set_position([box.x0, box.y0, box.width * 0.8, box.height])

    # Put a legend to the right of the current axis
    ax.legend(loc='center left', bbox_to_anchor=(1, 0.5))
    plt.grid(axis='y')
    if "electricityBus" in bus:
        plt.title("Monthly electricity balance for " + building.replace("__", ""))
    elif "spaceHeatingBus" in bus:
        plt.title("Monthly space heating balance for " + building.replace("__", ""))
    else:
        plt.title("Monthly domestic hot water balance for " + building.replace("__", ""))
    plt.show()


def hourly_daily_plot(data, bus, palette, new_legends):
    """
    Function for the bokeh plot of hourly and daily balance of a bus
    :param data: list of dict type, results from the optimization
    :param bus: list of str type, buses from which the summary is required
    :param palette: palette type form bokeh.palettes, different types can be found on
    https://docs.bokeh.org/en/latest/docs/reference/palettes.html or
    https://docs.bokeh.org/en/latest/_modules/bokeh/palettes.html
    :param new_legends: dict type, new legends to plot on the graph
    For example, Category10_8
    :return:
    """
    p_figs = []
    p_plots = []
    a = []
    b = 1

    for i in range(0, len(data)):
        l = []
        for j in bus[i]:
            l.append(j)
        l.reverse()
        l = l[0:11]  # in the case of a maximum of 9 buildings
        l.reverse()
        building = ''.join([str(item) for item in l])

        if "electricityBus" in bus[i]:
            dt = data[i]
            data_day = dt.resample('1d').sum()
            p1 = figure(title="Hourly electricity flows for " + building.replace("__", ""), x_axis_label="Date", y_axis_label="Power (kWh)", sizing_mode="scale_both")
            p2 = figure(title="Daily electricity flows for " + building.replace("__", ""), x_axis_label="Date", y_axis_label="Power (kWh)", sizing_mode="scale_both")
            colors = itertools.cycle(palette)
            for j, color in zip(dt.columns, colors):
                p1.line(dt.index, dt[j], legend_label=new_legends[j.replace(building, "")], color=color)
                p2.line(data_day.index, data_day[j], legend_label=new_legends[j.replace(building, "")], color=color)
            p_figs.append([p1, p2])
            p_plots.append(p1)
            p_plots.append(p2)

        elif "spaceHeatingBus" in bus[i]:
            dt = data[i]
            data_day = dt.resample('1d').sum()
            p3 = figure(title="Hourly space heating flows for " + building.replace("__", ""), x_axis_label="Date", y_axis_label="Power (kWh)", sizing_mode="scale_both")
            p4 = figure(title="Daily space heating flows for " + building.replace("__", ""), x_axis_label="Date", y_axis_label="Power (kWh)", sizing_mode="scale_both")
            colors = itertools.cycle(palette)
            for j, color in zip(dt.columns, colors):
                p3.line(dt.index, dt[j], legend_label=new_legends[j.replace(building, "")], color=color)
                p4.line(data_day.index, data_day[j], legend_label=new_legends[j.replace(building, "")], color=color)
            p_figs.append([p3, p4])
            p_plots.append(p3)
            p_plots.append(p4)

        else:
            dt = data[i]
            data_day = dt.resample('1d').sum()

            p5 = figure(title="Hourly domestic hot water flows for " + building.replace("__", ""), x_axis_label="Date", y_axis_label="Power (kWh)", sizing_mode="scale_both")
            p6 = figure(title="Daily domestic hot water flows for "  + building.replace("__", ""), x_axis_label="Date", y_axis_label="Power (kWh)", sizing_mode="scale_both")
            colors = itertools.cycle(palette)
            for j, color in zip(dt.columns, colors):
                p5.line(dt.index, dt[j], legend_label=new_legends[j.replace(building, "")], color=color)
                p6.line(data_day.index, data_day[j], legend_label=new_legends[j.replace(building, "")], color=color)
            p_figs.append([p5, p6])
            p_plots.append(p5)
            p_plots.append(p6)

    for p in p_plots:
        p.xaxis[0].formatter = DatetimeTickFormatter(months="%d %b")
        p.legend.click_policy = "hide"
    graph = layout(p_figs)

    html = file_html(graph, CDN, "bokeh_plots")
    file = open("bokeh_plots.html", 'w')
    file.write(html)
    file.close()

    show(graph)


def to_color(COLORS, obj=None):
    """
    Assign a deterministic pseudo-random color to argument.
    If COLORS[obj] is set, return that. For strings, this value depends only
    on the string content, so that same strings always yield the same color.
    :param COLORS: dict of components and their assigned color
    :param obj: any hashable object
    :return: a (r, g, b) color tuple if COLORS[obj] is set, otherwise a hexstring
    """
    try:
        color = tuple(rgb / 255.0 for rgb in COLORS[obj])
    except KeyError:
        # random deterministic color
        import hashlib
        color = '#' + hashlib.sha1(obj.encode()).hexdigest()[-6:]
    return color


def group_hbar_plots(ax, group_size, inner_sep=None):
    """
    Group bars of a horizontal barplot closer together.
    Given an existing horizontal bar plot handle ax, move bars of a given group size (>=2) closer together,
    reducing the distance within the bars of a group, but increasing the distance between different groups.
    By default, bars are placed within a coordinate system 1 unit apart. The space between two bars has
    size 1 - bar_height, which can be specified in matplotlib (and pandas) using the `width` argument.
    :param ax: matplotlib axis
    :param group_size: int type, how many bars to group together
    :param inner_sep: float type, vertical spacing within group (optional). Default: reduce the distance to a half
    :return:
    """
    handles, labels = ax.get_legend_handles_labels()
    bar_height = handles[0][0].get_height()  # assumption: identical for all

    if inner_sep is None:
        inner_sep = 0.5 * (1 - bar_height)

    for column, handle in enumerate(handles):
        for row, patch in enumerate(handle.patches):
            group_number, row_within_group = divmod(row, group_size)

            group_offset = (group_number * group_size
                            + 0.5 * (group_size - 1) * (1 - inner_sep)
                            - 0.5 * (group_size * bar_height))

            patch.set_y(row_within_group * (bar_height + inner_sep)
                        + group_offset)


def deduplicate_legend(handles, labels):
    """
    Remove double entries from figure legend.
    :param handles: list of legend entry handles
    :param labels: list of legend entry labels
    :return: (handles, labels) tuple of lists with duplicate labels removed
    """
    new_handles = []
    new_labels = []
    for hdl, lbl in zip(handles, labels):
        if not lbl in new_labels:
            new_handles.append(hdl)
            new_labels.append(lbl)
    # also, sort both lists accordingly
    new_labels, new_handles = (list(t) for t in zip(*sorted(zip(new_labels, new_handles))))
    return (new_handles, new_labels)


def resulting_data_diagram(elBus, shBus, dhwBus, costs, env, COLORS, building):
    """
    Function plotting the different results of the optimization. First, costs will be plotted, then the energy produced,
    comparing energy for electricity bus, sh and dhw bus, and finally the retrieved energy from the storages.
    :param elBus: dict type, results from the optimization applied to one bus.
    :param shBus: dict type, results from the optimization applied to one bus.
    :param dhwBus: dict type, results from the optimization applied to one bus.
    :param costs: dict type, resulting costs from the optimization
    :param env: dict type, resulting environmental impacts from the optimization
    :param COLORS: list type, different colors for the different components of the system
    :param building: str type, name of the building
    :return: Four bar plots and the costs, environmental impacts, production and storage dict created
    """
    production = {}
    storage = {}
    alpha = 0

    for flow in elBus.keys():
        if "Storage" in flow and "out" in new_legends[flow.replace("__Building"+str(building), "")]:
            storage[new_legends[flow.replace("__Building"+str(building), "")]] = [0, 0, sum(elBus[flow])]
        elif "Storage" not in flow and "Demand" not in flow:
            production[new_legends[flow.replace("__Building"+str(building), "")]] = [0, 0, sum(elBus[flow])]

            # specific for negative flows
            if "HP_SH" in flow or "HP_DHW" in flow:
                alpha += sum(elBus[flow])

    production["HP"] = [0, 0, -alpha]

    for flow in shBus.keys():
        if "Storage" in flow and "out" in new_legends[flow.replace("__Building"+str(building), "")]:
            storage[new_legends[flow.replace("__Building"+str(building), "")]] = [0, sum(shBus[flow]), 0]
        elif "Storage" not in flow and "Demand" not in flow:
            production[new_legends[flow.replace("__Building"+str(building), "")]] = [0, sum(shBus[flow]), 0]

    for flow in dhwBus.keys():
        if "Storage" in flow and "out" in new_legends[flow.replace("__Building"+str(building), "")]:
            storage[new_legends[flow.replace("__Building"+str(building), "")]] = [sum(dhwBus[flow]), 0, 0]
        elif "Storage" not in flow and "Demand" not in flow:
            production[new_legends[flow.replace("__Building"+str(building), "")]] = [sum(dhwBus[flow]), 0, 0]

    costs = costs.transpose()
    costs = costs.rename(index={0: building})

    env = env.transpose()
    for i in env.columns:
        for j in new_legends.keys():
            if i.replace("__Building"+str(building), "") in j:
                env = env.rename(columns={i: new_legends[j]})

    production = pd.DataFrame.from_dict(production, orient='index')
    production = production.transpose()

    storage = pd.DataFrame.from_dict(storage, orient='index', columns=["dhw", "sh", "elec"])
    storage = storage.transpose()

    fig = plt.figure(figsize=(18, 8))
    gs = gridspec.GridSpec(1, 4, width_ratios=[3, 4, 8, 3], wspace=0.03)

    ax0 = plt.subplot(gs[0])
    costs_colors = [to_color(COLORS, x) for x in costs.columns]
    bp0 = costs.plot(ax=ax0, kind='barh', color=costs_colors, stacked=True, linewidth=0)

    ax1 = plt.subplot(gs[1])
    env_colors = [to_color(COLORS, x) for x in env.columns]
    bp1 = env.plot(ax=ax1, kind='barh', color=env_colors, stacked=True, linewidth=0)

    ax2 = plt.subplot(gs[2])
    production_colors = [to_color(COLORS, x) for x in production.columns]
    bp2 = production.plot(ax=ax2, kind='barh', color=production_colors, stacked=True, linewidth=0, width=.5)

    ax3 = plt.subplot(gs[3])
    storage_colors = [to_color(COLORS, x) for x in storage.columns]
    bp3 = storage.plot(ax=ax3, kind='barh', color=storage_colors, stacked=True, linewidth=0)

    # remove scenario names from other bar plots
    for ax in [ax1, ax2]:
        ax.set_yticklabels('')
    for ax in [ax2, ax3]:
        group_hbar_plots(ax, group_size=3, inner_sep=0.01)
    ax3.yaxis.tick_right()

    # set limits and ticks for both axes
    for ax in [ax0, ax1, ax2, ax3]:
        ax.yaxis.grid(False)
        ax.xaxis.grid(True, 'major', linestyle='-')
        ax.xaxis.set_ticks_position('none')
        ax.yaxis.set_ticks_position('none')

        # group 1,000,000 with commas
        xmin, xmax = ax.get_xlim()
        if xmax > 90 or xmin < -90:
            group_thousands_and_skip_first = tkr.FuncFormatter(
                lambda x, pos: '' if pos == 0 else '{:0,d}'.format(int(x)))
            ax.xaxis.set_major_formatter(group_thousands_and_skip_first)
        else:
            skip_lowest = tkr.FuncFormatter(
                lambda x, pos: '' if pos == 0 else x)
            ax.xaxis.set_major_formatter(skip_lowest)

        # legend
        # set style arguments
        legend_style = {'frameon': False,
                        'loc': 'lower center',
                        'ncol': 2,
                        'bbox_to_anchor': (0.5, .99)}
        # get handels and labels, remove duplicate labels
        handles, labels = deduplicate_legend(*ax.get_legend_handles_labels())
        # set legend to use those
        lg = ax.legend(handles, labels, **legend_style)
        # finally, remove lines from patches
        plt.setp(lg.get_patches(), linewidth=0)

    ax0.set_xlabel('Total costs (CHF)')
    ax1.set_xlabel('Total environmental impacts (kgCo2eq)')
    ax2.set_xlabel('Total energy produced (kWh)')
    ax3.set_xlabel('Retrieved energy (kWh)')
    return fig, costs, env, production, storage


def resulting_data_diagram_loop(elec, sh, dhw, costs, env, colors, buildings):
    """
    Function plotting the graph comparing the different buildings/scenarios on costs, energy produced and energy
    retrieved from storages
    :param elec: list of dict type, optimization results
    :param sh: list of dict type, optimization results
    :param dhw: list of dict type, optimization results
    :param costs: list of dict type, optimization results
    :param env: list of dict type, optimization results
    :param colors: list type, different colors for the different components of the system
    :param buildings: list of str type, name of the different buildings
    :return: figure created
    """
    n_costs, n_env, n_production, n_storage = resulting_data_diagram(elec[0], sh[0], dhw[0], costs[0], env[0], colors, buildings[0])[1:]
    for i in range(1, len(elec)):
        a, b, c, d = resulting_data_diagram(elec[i], sh[i], dhw[i], costs[i], env[i], colors, buildings[i])[1:]
        n_costs = pd.concat([n_costs, a])
        n_env = pd.concat([n_env, b])
        n_production = pd.concat([n_production, c])
        n_storage = pd.concat([n_storage, d])

    fig = plt.figure(figsize=(18, 8))
    gs = gridspec.GridSpec(1, 4, width_ratios=[4, 3, 8, 3], wspace=0.03)

    ax0 = plt.subplot(gs[0])
    costs_colors = [to_color(colors, x) for x in n_costs.columns]
    bp0 = n_costs.plot(ax=ax0, kind='barh', color=costs_colors, stacked=True, linewidth=0)

    ax1 = plt.subplot(gs[1])
    env_colors = [to_color(colors, x) for x in n_env.columns]
    bp1 = n_env.plot(ax=ax1, kind='barh', color=env_colors, stacked=True, linewidth=0)

    ax2 = plt.subplot(gs[2])
    production_colors = [to_color(colors, x) for x in n_production.columns]
    bp2 = n_production.plot(ax=ax2, kind='barh', color=production_colors, stacked=True, linewidth=0, width=.5)

    ax3 = plt.subplot(gs[3])
    storage_colors = [to_color(colors, x) for x in n_storage.columns]
    bp3 = n_storage.plot(ax=ax3, kind='barh', color=storage_colors, stacked=True, linewidth=0)

    # remove scenario names from other bar plots
    for ax in [ax1, ax2, ax3]:
        ax.set_yticklabels('')
    for ax in [ax2, ax3]:
        group_hbar_plots(ax, group_size=3, inner_sep=0.01)

    # set limits and ticks for both axes
    for ax in [ax0, ax1, ax2, ax3]:
        ax.yaxis.grid(False)
        ax.xaxis.grid(True, 'major', linestyle='-')
        ax.xaxis.set_ticks_position('none')
        ax.yaxis.set_ticks_position('none')

        # group 1,000,000 with commas
        xmin, xmax = ax.get_xlim()
        if xmax > 90 or xmin < -90:
            group_thousands_and_skip_first = tkr.FuncFormatter(
                lambda x, pos: '' if pos == 0 else '{:0,d}'.format(int(x)))
            ax.xaxis.set_major_formatter(group_thousands_and_skip_first)
        else:
            skip_lowest = tkr.FuncFormatter(
                lambda x, pos: '' if pos == 0 else x)
            ax.xaxis.set_major_formatter(skip_lowest)

        # legend
        # set style arguments
        legend_style = {'frameon': False,
                        'loc': 'lower center',
                        'ncol': 2,
                        'bbox_to_anchor': (0.5, .99)}
        # get handels and labels, remove duplicate labels
        handles, labels = deduplicate_legend(*ax.get_legend_handles_labels())
        # set legend to use those
        lg = ax.legend(handles, labels, **legend_style)
        # finally, remove lines from patches
        plt.setp(lg.get_patches(), linewidth=0)

    ax0.set_xlabel('Total costs (CHF)')
    ax1.set_xlabel('Total environmental impacts (kgCo2eq)')
    ax2.set_xlabel('Total energy produced (kWh)\n First line: electricity, Second line: space heating,\n Third line: domestic hot water')
    ax3.set_xlabel('Retrieved energy (kWh)\n First line: electricity, Second line: space heating,\n Third line: domestic hot water')

    return fig

def resulting_data_demand_diagram(elBus, shBus, dhwBus, COLORS, building):
    """
    Function plotting the different results of the optimization. First, costs will be plotted, then the energy produced,
    comparing energy for electricity bus, sh and dhw bus, and finally the retrieved energy from the storages.
    :param elBus: dict type, results from the optimization applied to one bus. Called like "solph.views.node(results, bus)"
    :param shBus: dict type, results from the optimization applied to one bus. Called like "solph.views.node(results, bus)"
    :param dhwBus: dict type, results from the optimization applied to one bus. Called like "solph.views.node(results, bus)"
    :param COLORS: list type, different colors for the different components of the system
    :param building: str type, name of the building
    :return: Three bar plots and the elec, sh and dhw dict created
    """
    elec = {}
    sh = {}
    dhw = {}
    for flow in elBus.keys():
        if "Demand" in flow:
            elec[new_legends[flow.replace("__Building"+str(building), "")]] = [0, sum(elBus[flow])]
        else:
            elec[new_legends[flow.replace("__Building"+str(building), "")]] = [sum(elBus[flow]), 0]

        if "Storage" in flow and "in" in new_legends[flow.replace("__Building"+str(building), "")]:
            del elec[new_legends[flow.replace("__Building"+str(building), "")]]
        elif "electricityBus" not in flow[5:] and "Demand" not in flow:
            del elec[new_legends[flow.replace("__Building" + str(building), "")]]
        elif "Resource" in flow or "excess" in flow:
            del elec[new_legends[flow.replace("__Building" + str(building), "")]]

    for flow in shBus.keys():
        if "Demand" in flow:
            sh[new_legends[flow.replace("__Building" + str(building), "")]] = [0, sum(shBus[flow])]
        else:
            sh[new_legends[flow.replace("__Building" + str(building), "")]] = [sum(shBus[flow]), 0]
        if "Storage" in flow and "in" in new_legends[flow.replace("__Building"+str(building), "")]:
            del sh[new_legends[flow.replace("__Building"+str(building), "")]]

    for flow in dhwBus.keys():
        if "Demand" in flow:
            dhw[new_legends[flow.replace("__Building" + str(building), "")]] = [0, sum(dhwBus[flow])]
        else:
            dhw[new_legends[flow.replace("__Building" + str(building), "")]] = [sum(dhwBus[flow]), 0]

        if "Storage" in flow and "in" in new_legends[flow.replace("__Building"+str(building), "")]:
            del dhw[new_legends[flow.replace("__Building"+str(building), "")]]

    elec = pd.DataFrame.from_dict(elec, orient='index')
    elec = elec.transpose()

    sh = pd.DataFrame.from_dict(sh, orient='index')
    sh = sh.transpose()

    dhw = pd.DataFrame.from_dict(dhw, orient='index')
    dhw = dhw.transpose()

    fig = plt.figure(figsize=(18, 8))
    gs = gridspec.GridSpec(1, 3, wspace=0.1)

    ax0 = plt.subplot(gs[0])
    elec_colors = [to_color(COLORS, x) for x in elec.columns]
    bp0 = elec.plot(ax=ax0, kind='barh', color=elec_colors, stacked=True, linewidth=0, width=.5)

    ax1 = plt.subplot(gs[1])
    sh_colors = [to_color(COLORS, x) for x in sh.columns]
    bp1 = sh.plot(ax=ax1, kind='barh', color=sh_colors, stacked=True, linewidth=0)

    ax2 = plt.subplot(gs[2])
    dhw_colors = [to_color(COLORS, x) for x in dhw.columns]
    bp2 = dhw.plot(ax=ax2, kind='barh', color=dhw_colors, stacked=True, linewidth=0)

    # remove scenario names from other bar plots
    ax0.set_yticklabels(('', building))
    for ax in [ax1, ax2]:
        ax.set_yticklabels('')
    for ax in [ax0, ax1, ax2]:
        group_hbar_plots(ax, group_size=2, inner_sep=0.0)

    # set limits and ticks for both axes
    for ax in [ax0, ax1, ax2]:
        ax.yaxis.grid(False)
        ax.xaxis.grid(True, 'major', linestyle='-')
        ax.xaxis.set_ticks_position('none')
        ax.yaxis.set_ticks_position('none')

        # group 1,000,000 with commas
        xmin, xmax = ax.get_xlim()
        if xmax > 90 or xmin < -90:
            group_thousands_and_skip_first = tkr.FuncFormatter(
                lambda x, pos: '' if pos == 0 else '{:0,d}'.format(int(x)))
            ax.xaxis.set_major_formatter(group_thousands_and_skip_first)
        else:
            skip_lowest = tkr.FuncFormatter(
                lambda x, pos: '' if pos == 0 else x)
            ax.xaxis.set_major_formatter(skip_lowest)

        # legend
        # set style arguments
        legend_style = {'frameon': False,
                        'loc': 'lower center',
                        'ncol': 3,
                        'bbox_to_anchor': (0.5, .99)}
        # get handels and labels, remove duplicate labels
        handles, labels = deduplicate_legend(*ax.get_legend_handles_labels())
        # set legend to use those
        lg = ax.legend(handles, labels, **legend_style)
        # finally, remove lines from patches
        plt.setp(lg.get_patches(), linewidth=0)

    ax0.set_xlabel('Demand and energy produced\n for electricity (kWh)')
    ax1.set_xlabel('Demand and energy produced\n for space heat (kWh)')
    ax2.set_xlabel('Demand and energy produced\n for domestic hot water (kWh)')

    return fig, elec, sh, dhw

def resulting_data_demand_diagram_loop(elec, sh, dhw, colors, buildings):
    """
    Function plotting the graph comparing the different buildings/scenarios on costs, energy produced and energy
    retrieved from storages
    :param elec: list of dict type, optimization results
    :param sh: list of dict type, optimization results
    :param dhw: list of dict type, optimization results
    :param colors: list type, different colors for the different components of the system
    :param buildings: list of str type, name of the different buildings
    :return: figure created
    """
    n_elec, n_sh, n_dhw = resulting_data_demand_diagram(elec[0], sh[0], dhw[0], colors, int(buildings[0]))[1:]
    for i in range(1, len(elec)):
        a, b, c = resulting_data_demand_diagram(elec[i], sh[i], dhw[i], colors, buildings[i])[1:]
        n_elec = pd.concat([n_elec, a])
        n_sh = pd.concat([n_sh, b])
        n_dhw = pd.concat([n_dhw, c])

    fig = plt.figure(figsize=(18, 8))
    gs = gridspec.GridSpec(1, 3, wspace=0.1)

    ax0 = plt.subplot(gs[0])
    elec_colors = [to_color(colors, x) for x in n_elec.columns]
    bp0 = n_elec.plot(ax=ax0, kind='barh', color=elec_colors, stacked=True, linewidth=0, width=.5)

    ax1 = plt.subplot(gs[1])
    sh_colors = [to_color(colors, x) for x in n_sh.columns]
    bp1 = n_sh.plot(ax=ax1, kind='barh', color=sh_colors, stacked=True, linewidth=0)

    ax2 = plt.subplot(gs[2])
    dhw_colors = [to_color(colors, x) for x in n_dhw.columns]
    bp2 = n_dhw.plot(ax=ax2, kind='barh', color=dhw_colors, stacked=True, linewidth=0)

    # remove scenario names from other bar plots
    liste = []
    for i in buildings:
        liste.append('')
        liste.append(i)
    ax0.set_yticklabels(liste)
    for ax in [ax1, ax2]:
        ax.set_yticklabels('')
    for ax in [ax0, ax1, ax2]:
        group_hbar_plots(ax, group_size=2, inner_sep=0.0)

    # set limits and ticks for both axes
    for ax in [ax0, ax1, ax2]:
        ax.yaxis.grid(False)
        ax.xaxis.grid(True, 'major', linestyle='-')
        ax.xaxis.set_ticks_position('none')
        ax.yaxis.set_ticks_position('none')

        # group 1,000,000 with commas
        xmin, xmax = ax.get_xlim()
        if xmax > 90 or xmin < -90:
            group_thousands_and_skip_first = tkr.FuncFormatter(
                lambda x, pos: '' if pos == 0 else '{:0,d}'.format(int(x)))
            ax.xaxis.set_major_formatter(group_thousands_and_skip_first)
        else:
            skip_lowest = tkr.FuncFormatter(
                lambda x, pos: '' if pos == 0 else x)
            ax.xaxis.set_major_formatter(skip_lowest)

        # legend
        # set style arguments
        legend_style = {'frameon': False,
                        'loc': 'lower center',
                        'ncol': 3,
                        'bbox_to_anchor': (0.5, .99)}
        # get handels and labels, remove duplicate labels
        handles, labels = deduplicate_legend(*ax.get_legend_handles_labels())
        # set legend to use those
        lg = ax.legend(handles, labels, **legend_style)
        # finally, remove lines from patches
        plt.setp(lg.get_patches(), linewidth=0)

    ax0.set_xlabel('Demand and energy produced\n for electricity (kWh)')
    ax1.set_xlabel('Demand and energy produced\n for space heat (kWh)')
    ax2.set_xlabel('Demand and energy produced\n for domestic hot water (kWh)')

    return fig


def get_data(filepath):
    """
    Function for the results recovery from an Excel file
    :param filepath: path to the Excel containing the results of the optimization
    :return: the different dicts created during the optimization
    """
    wb = load_workbook(filepath, read_only=True, keep_links=False)
    dict_sheet = {}
    for sheet in wb.sheetnames:
        dict_sheet[sheet] = pd.read_excel(filepath, sheet_name=sheet, index_col=0, engine='openpyxl')

    return dict_sheet


if __name__ == '__main__':
    #####################################
    ########## Classic plots  ###########
    #####################################

    new_legends = {
        "(('naturalGasResource', 'naturalGasBus'), 'flow')": "NaturalGas",
        "(('electricityBus', 'excesselectricityBus'), 'flow')": "Feed-in",
        "(('electricityBus', 'electricalStorage'), 'flow')": "Battery_in",
        "(('electricityBus', 'electricityDemand'), 'flow')": "Demand_elec",
        "(('electricityBus', 'HP_DHW'), 'flow')": "HP_dhw",
        "(('electricityBus', 'HP_SH'), 'flow')": "HP_sh",
        "(('CHP', 'electricityBus'), 'flow')": "CHP_elec",
        "(('electricalStorage', 'electricityBus'), 'flow')": "Battery_out",
        "(('electricityResource', 'electricityBus'), 'flow')": "Grid_purchase",
        "(('dhwStorage', 'domesticHotWaterBus'), 'flow')": "Storage_dhw_out",
        "(('domesticHotWaterBus', 'dhwStorage'), 'flow')": "Storage_dhw_in",
        "(('domesticHotWaterBus', 'domesticHotWaterDemand'), 'flow')": "Demand_dhw",
        "(('HP_DHW', 'domesticHotWaterBus'), 'flow')": "HP_dhw",
        "(('CHP', 'domesticHotWaterBus'), 'flow')": "CHP_dhw",
        "(('shStorage', 'spaceHeatingBus'), 'flow')": "Storage_sh_out",
        "(('spaceHeatingBus', 'shStorage'), 'flow')": "Storage_sh_in",
        "(('spaceHeatingBus', 'spaceHeatingDemand'), 'flow')": "Demand_sh",
        "(('CHP', 'spaceHeatingBus'), 'flow')": "CHP_sh",
        "(('HP_SH', 'spaceHeatingBus'), 'flow')": "HP_sh",
    }

    buses = get_data("results.xlsx")
    elec_names = []
    elec_dict = []
    sh_names = []
    sh_dict = []
    dhw_names = []
    dhw_dict = []
    costs_names = []
    costs_dict = []
    env_names = []
    env_dict = []

    buildings_dict = []
    buildings_names = []
    buildings_number = []
    beta = list(buses.keys())
    for i in beta:
        alpha = []
        alpha_a = []
        l = []
        for k in i:
            l.append(k)
        building = l[-1]  # in the case of a maximum of 9 buildings
        beta_bis = beta.copy()
        for j in beta:
            if j.endswith(building) and ("electricityBus" in j or "spaceHeatingBus" in j or "domesticHotWaterBus" in j):
                alpha.append(j)
                alpha_a.append(buses[j])
                beta_bis.remove(j)
        if alpha != []:
            buildings_dict.append(alpha_a)
            buildings_names.append(alpha)
            buildings_number.append(building)
        for a in alpha:
            beta.remove(a)

    for i in buses.keys():
        if "electricityBus" in i:
            elec_names.append(i)
            elec_dict.append(buses[i])
        if "spaceHeatingBus" in i:
            sh_names.append(i)
            sh_dict.append(buses[i])
        if "domesticHotWater" in i:
            dhw_names.append(i)
            dhw_dict.append(buses[i])
        if "costs" in i:
            costs_names.append(i)
            costs_dict.append(buses[i])
        if "env_impacts" in i:
            env_names.append(i)
            env_dict.append(buses[i])

    for i in elec_names + sh_names + dhw_names:
        monthly_balance(buses[i], i, new_legends)

    hourly_daily_plot(elec_dict, elec_names, Category10_8, new_legends)
    hourly_daily_plot(sh_dict, sh_names, Category10_8, new_legends)
    hourly_daily_plot(dhw_dict, dhw_names, Category10_8, new_legends)

    for i in range(len(buildings_names)):
        hourly_daily_plot(buildings_dict[i], buildings_names[i], Category10_8, new_legends)

    #####################################
    ## Summary of the whole experiment ##
    #####################################

    my_colors = {
        'Battery_out': (100, 160, 200),
        'Storage_sh_out': (196, 7, 27),
        'Storage_dhw_out': (196, 7, 27),
        'HP_dhw': (62, 173, 0),
        'HP_sh': (62, 173, 0),
        'HP': (62, 173, 0),
        'CHP_elec': (0, 101, 189),
        'CHP_sh': (0, 101, 189),
        'CHP_dhw': (0, 101, 189),
        'CHP': (0, 101, 189),
        'Inputs': (252, 93, 93),
        'Operation': (252, 93, 93),
        'Investment': (0, 119, 138),
        'Feed-in': (218, 215, 203),
        'Demand_elec': (131, 166, 151),
        'Demand_sh': (131, 166, 151),
        'Demand_dhw': (131, 166, 151),
        'Grid_purchase': (237, 127, 16),
    }
    COLORS = {}
    for name, color in my_colors.items():
        COLORS[name] = color

    for i in range(len(buildings_names)):
        fig1 = resulting_data_diagram(elec_dict[i], sh_dict[i], dhw_dict[i], costs_dict[i], env_dict[i], COLORS, buildings_number[i])[0]
        fig2 = resulting_data_demand_diagram(elec_dict[i], sh_dict[i], dhw_dict[i], COLORS, buildings_number[i])[0]

    fig3 = resulting_data_diagram_loop(elec_dict, sh_dict, dhw_dict, costs_dict, env_dict, COLORS, buildings_number)

    fig4 = resulting_data_demand_diagram_loop(elec_dict, sh_dict, dhw_dict, COLORS, buildings_number)

    plt.show()

logging.info("Done!")

