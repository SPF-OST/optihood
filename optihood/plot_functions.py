import matplotlib.pyplot as plt
import matplotlib.gridspec as gridspec
import matplotlib.ticker as tkr

from bokeh.plotting import figure, show
from bokeh.layouts import layout, gridplot
from bokeh.models import DatetimeTickFormatter, HoverTool, Legend
from bokeh.palettes import *
from bokeh.io import output_file

from openpyxl import load_workbook
from datetime import datetime
from dateutil.parser import isoparse
import itertools
import pandas as pd
import os

# This file defines different functions for the plotting of the results of the optimization.
# The plots are made at the end of this file, introducing a .xls file previously created during the optimization.
# An important parameter "optMode" needs to be set when running the code (look in the __main__ fragment)


def monthlyBalance(data, bus, new_legends):
    """
    Function for the definition of the monthly summary of a bus
    :param data: dict type, results from the optimization applied to one bus
    :param bus: str type, bus from which the summary is required
    :param new_legends: dict type, new legends to plot on the graph
    :return:
    """
    building = "__" + bus.split("__")[1]
    if not data.filter(like='electricityLink').empty:
        elLinks = data.filter(like='electricityLink')  # half would be el_link_out and half would be el_link_in
        data.drop(list(data.filter(regex='electricityLink')), axis=1, inplace=True)
        mid = int(len(elLinks.columns) / 2)
        elLinksOut = elLinks.iloc[:, 0:mid]
        elLinksIn = elLinks.iloc[:, mid:len(elLinks.columns)]
        elLinksOut["(('electricityBus', 'electricityLink'), 'flow')"] = elLinksOut.sum(axis=1)
        elLinksIn["(('electricityLink', 'electricityInBus'), 'flow')"] = elLinksIn.sum(axis=1)
        data = pd.concat((data, elLinksIn["(('electricityLink', 'electricityInBus'), 'flow')"]), axis=1)
        data = pd.concat((data, elLinksOut["(('electricityBus', 'electricityLink'), 'flow')"]), axis=1)
    data_month = data.resample('1m').sum()
    monthShortNames = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']

    while len(data_month) != len(monthShortNames):
        data.drop(data.index[-1], inplace=True)
        data_month = data.resample('1m').sum()
    neg_flow = []
    pos_flow = []
    for i in data.columns:
        a = [i.strip("()").split(", ")]
        if "Bus" in a[0][0]:
            neg_flow.append(i)
        else:
            pos_flow.append(i)
    plt.figure()
    mark = []
    for i in neg_flow:
        plt.bar(monthShortNames, -data_month[i], label=new_legends[i.replace(building, "")], bottom=sum(mark))
        mark.append(-data_month[i])
    mark = []
    for i in pos_flow:
        plt.bar(monthShortNames, data_month[i], label=new_legends[i.replace(building, "")], bottom=sum(mark))
        mark.append(data_month[i])

    ax = plt.gca()
    # Shrink current axis by 20%
    box = ax.get_position()
    ax.set_position([box.x0, box.y0, box.width * 0.8, box.height])

    # Put a legend to the right of the current axis
    ax.legend(loc='center left', bbox_to_anchor=(1, 0.5))
    plt.grid(axis='y')
    if "electricity" in bus or "grid" in bus:
        plt.title("Monthly electricity balance for " + building.replace("__", ""))
    elif "spaceHeating" in bus:
        plt.title("Monthly space heating balance for " + building.replace("__", ""))
    else:
        plt.title("Monthly domestic hot water balance for " + building.replace("__", ""))
    plt.show()


def hourlyDailyPlot(data, bus, palette, new_legends):
    """
    Function for the bokeh plot of hourly and daily balance of a bus
    :param data: list of dict type, results from the optimization
    :param bus: list of str type, buses from which the summary is required
    :param palette: palette type form bokeh.palettes (For example, Category10_8), different types can be found on
    https://docs.bokeh.org/en/latest/docs/reference/palettes.html or
    https://docs.bokeh.org/en/latest/_modules/bokeh/palettes.html
    :param new_legends: dict type, new legends to plot on the graph
    For example, Category10_8
    :return:
    """
    p_figs = []
    p_figs_h = []
    p_figs_d = []
    p_plots = []
    a = []
    b = 1

    for i in range(0, len(data)):
        building = '__' + bus[i].split("__")[1]

        if "electricity" in bus[i]:
            del bus[i + 1]
            del bus[i + 1]
            dt = data[i]
            dt.pop(f"(('electricityProdBus{building}', 'electricitySource{building}'), 'flow')")
            elLinks = dt.filter(like='electricityLink')  #half would be el_link_out and half would be el_link_in
            dt.drop(list(dt.filter(regex = 'electricityLink')), axis = 1, inplace = True)
            mid = int(len(elLinks.columns)/2)
            elLinksOut = elLinks.iloc[:, 0:mid]
            elLinksIn = elLinks.iloc[:, mid:len(elLinks.columns)]
            elLinksOut["(('electricityBus', 'electricityLink'), 'flow')"] = elLinksOut.sum(axis=1)
            elLinksIn["(('electricityLink', 'electricityInBus'), 'flow')"] = elLinksIn.sum(axis=1)
            dt = pd.concat((dt, elLinksIn["(('electricityLink', 'electricityInBus'), 'flow')"]), axis=1)
            dt = pd.concat((dt, elLinksOut["(('electricityBus', 'electricityLink'), 'flow')"]), axis=1)

            data_day = dt.resample('1d').sum()
            p1 = figure(title="Hourly electricity flows for " + building.replace("__", ""), x_axis_label="Date", y_axis_label="Energy (kWh)", sizing_mode="scale_both")
            p1.add_layout(Legend(), 'right')
            p1.add_tools(HoverTool(tooltips=[('Time', '@x{%d/%m/%Y %H:%M:%S}'), ('Energy', '@y{0.00}')],
                                   formatters={'@x': 'datetime'},
                                   mode='mouse'))
            p2 = figure(title="Daily electricity flows for " + building.replace("__", ""), x_axis_label="Date", y_axis_label="Energy (kWh)", sizing_mode="scale_both")
            p2.add_layout(Legend(), 'right')
            p2.add_tools(HoverTool(tooltips=[('Date', '@x{%d/%m/%Y}'), ('Energy', '@y{0.00}')],
                                   formatters={'@x': 'datetime'},
                                   mode='mouse'))
            if len(p_figs_h) > 0:
                p1.x_range = p_figs_h[0].x_range
                p2.x_range = p_figs_d[0].x_range
            colors = itertools.cycle(palette)
            for j, color in zip(dt.columns, colors):
                p1.line(dt.index, dt[j], legend_label=new_legends[j.replace(building, "")], color=color, line_width=1.5)
                p2.line(data_day.index, data_day[j], legend_label=new_legends[j.replace(building, "")], color=color, line_width=1.5)
            p_figs.append([p1, p2])
            p_figs_h.append(p1)
            p_figs_d.append(p2)
            p_plots.append(p1)
            p_plots.append(p2)

        elif "shSource" in bus[i]:
            dt = data[i]
            shLinks = dt.filter(like='shLink')  # half would be el_link_out and half would be el_link_in
            dt.drop(list(dt.filter(regex='shLink')), axis=1, inplace=True)
            mid = int(len(shLinks.columns) / 2)
            shLinksOut = shLinks.iloc[:, 0:mid]
            shLinksIn = shLinks.iloc[:, mid:len(shLinks.columns)]
            shLinksOut["(('spaceHeatingBus', 'shLink'), 'flow')"] = shLinksOut.sum(axis=1)
            shLinksIn["(('shLink', 'shDemandBus'), 'flow')"] = shLinksIn.sum(axis=1)
            dt = pd.concat((dt, shLinksIn["(('shLink', 'shDemandBus'), 'flow')"]), axis=1)
            dt = pd.concat((dt, shLinksOut["(('spaceHeatingBus', 'shLink'), 'flow')"]), axis=1)
            data_day = dt.resample('1d').sum()
            p3 = figure(title="Hourly space heating flows for " + building.replace("__", ""), x_axis_label="Date", y_axis_label="Energy (kWh)", sizing_mode="scale_both")
            p3.add_layout(Legend(), 'right')
            p3.add_tools(HoverTool(tooltips=[('Time', '@x{%d/%m/%Y %H:%M:%S}'), ('Energy', '@y{0.00}')],
                                   formatters={'@x': 'datetime'},
                                   mode='mouse'))
            p4 = figure(title="Daily space heating flows for " + building.replace("__", ""), x_axis_label="Date", y_axis_label="Energy (kWh)", sizing_mode="scale_both")
            p4.add_layout(Legend(), 'right')
            p4.add_tools(HoverTool(tooltips=[('Date', '@x{%d/%m/%Y}'), ('Energy', '@y{0.00}')],
                                   formatters={'@x': 'datetime'},
                                   mode='mouse'))
            if len(p_figs_h) > 0:
                p3.x_range = p_figs_h[0].x_range
                p4.x_range = p_figs_d[0].x_range
            colors = itertools.cycle(palette)
            for j, color in zip(dt.columns, colors):
                p3.line(dt.index, dt[j], legend_label=new_legends[j.replace(building, "")], color=color, line_width=1.5)
                p4.line(data_day.index, data_day[j], legend_label=new_legends[j.replace(building, "")], color=color, line_width=1.5)
            p_figs.append([p3, p4])
            p_figs_h.append(p3)
            p_figs_d.append(p4)
            p_plots.append(p3)
            p_plots.append(p4)

        else:
            dt = data[i]
            data_day = dt.resample('1d').sum()

            p5 = figure(title="Hourly domestic hot water flows for " + building.replace("__", ""), x_axis_label="Date", y_axis_label="Energy (kWh)", sizing_mode="scale_both")
            p5.add_layout(Legend(), 'right')
            p5.add_tools(HoverTool(tooltips=[('Time', '@x{%d/%m/%Y %H:%M:%S}'), ('Energy', '@y{0.00}')],
                                   formatters={'@x': 'datetime'},
                                   mode='mouse'))
            p6 = figure(title="Daily domestic hot water flows for "  + building.replace("__", ""), x_axis_label="Date", y_axis_label="Energy (kWh)", sizing_mode="scale_both")
            p6.add_layout(Legend(), 'right')
            p6.add_tools(HoverTool(tooltips=[('Date', '@x{%d/%m/%Y}'), ('Energy', '@y{0.00}')],
                                   formatters={'@x': 'datetime'},
                                   mode='mouse'))
            if len(p_figs_h) > 0:
                p5.x_range = p_figs_h[0].x_range
                p6.x_range = p_figs_d[0].x_range
            colors = itertools.cycle(palette)
            for j, color in zip(dt.columns, colors):
                p5.line(dt.index, dt[j], legend_label=new_legends[j.replace(building, "")], color=color, line_width=1.5)
                p6.line(data_day.index, data_day[j], legend_label=new_legends[j.replace(building, "")], color=color, line_width=1.5)
            p_figs.append([p5, p6])
            p_figs_h.append(p5)
            p_figs_d.append(p6)
            p_plots.append(p5)
            p_plots.append(p6)

    for p in p_plots:
        p.xaxis[0].formatter = DatetimeTickFormatter(months="%d %b")
        p.legend.click_policy = "hide"
        p.legend.location = "top_left"
        p.legend.label_text_font_size = "8pt"

    return (p_figs_h, p_figs_d)


def toColor(COLORS, obj=None):
    """
    Function from the URBS platform https://github.com/ojdo/urbs/blob/1house/comp.py
    Assign a deterministic pseudo-random color to argument.
    If COLORS[obj] is set, return that. For strings, this value depends only
    on the string content, so that same strings always yield the same color.
    :param COLORS: dict of components and their assigned color
    :param obj: any hashable object
    :return: a (r, g, b) color tuple if COLORS[obj] is set, otherwise a hexstring
    """
    try:
        color = tuple(rgb / 255.0 for rgb in COLORS[obj])
    except KeyError:
        # random deterministic color
        import hashlib
        color = '#' + hashlib.sha1(obj.encode()).hexdigest()[-6:]
    return color


def groupHbarPlots(ax, group_size, inner_sep=None):
    """
    Function from the URBS platform https://github.com/ojdo/urbs/blob/1house/comp.py
    Group bars of a horizontal barplot closer together.
    Given an existing horizontal bar plot handle ax, move bars of a given group size (>=2) closer together,
    reducing the distance within the bars of a group, but increasing the distance between different groups.
    By default, bars are placed within a coordinate system 1 unit apart. The space between two bars has
    size 1 - bar_height, which can be specified in matplotlib (and pandas) using the `width` argument.
    :param ax: matplotlib axis
    :param group_size: int type, how many bars to group together
    :param inner_sep: float type, vertical spacing within group (optional). Default: reduce the distance to a half
    :return:
    """
    handles, labels = ax.get_legend_handles_labels()
    bar_height = handles[0][0].get_height()  # assumption: identical for all

    if inner_sep is None:
        inner_sep = 0.5 * (1 - bar_height)

    for column, handle in enumerate(handles):
        for row, patch in enumerate(handle.patches):
            group_number, row_within_group = divmod(row, group_size)

            group_offset = (group_number * group_size
                            + 0.5 * (group_size - 1) * (1 - inner_sep)
                            - 0.5 * (group_size * bar_height))

            patch.set_y(row_within_group * (bar_height + inner_sep)
                        + group_offset)


def deduplicateLegend(handles, labels):
    """
    Function from the URBS platform https://github.com/ojdo/urbs/blob/1house/comp.py
    Remove double entries from figure legend.
    :param handles: list of legend entry handles
    :param labels: list of legend entry labels
    :return: (handles, labels) tuple of lists with duplicate labels removed
    """
    new_handles = []
    new_labels = []
    for hdl, lbl in zip(handles, labels):
        if not lbl in new_labels:
            new_handles.append(hdl)
            new_labels.append(lbl)
    # also, sort both lists accordingly
    new_labels, new_handles = (list(t) for t in zip(*sorted(zip(new_labels, new_handles))))
    return (new_handles, new_labels)


def resultingDataDiagram(elBus, shBus, dhwBus, costs, env, COLORS, building, newLegends):
    """
    Function inspired from the URBS platform https://github.com/ojdo/urbs/blob/1house/comp.py
    Function plotting the different results of the optimization. First, costs will be plotted, then the energy produced,
    comparing energy for electricity bus, sh and dhw bus, and finally the retrieved energy from the storages.
    :param elBus: dict type, results from the optimization applied to one bus.
    :param shBus: dict type, results from the optimization applied to one bus.
    :param dhwBus: dict type, results from the optimization applied to one bus.
    :param costs: dict type, resulting costs from the optimization
    :param env: dict type, resulting environmental impacts from the optimization
    :param COLORS: list type, different colors for the different components of the system
    :param building: str type, name of the building
    :return: Four bar plots and the costs, environmental impacts, production and storage dict created
    """
    production = {}
    storage = {}
    alpha = 0
    if optMode == "group":
        list = []

    for flow in elBus.keys():
        if "Storage" in flow and "out" in newLegends[flow.replace("__Building"+str(building), "")]:
            storage[newLegends[flow.replace("__Building"+str(building), "")]] = [0, 0, sum(elBus[flow])]
        elif optMode == "group" and "electricityLink" in flow:
            if "_in" in newLegends[flow.replace("__Building" + str(building), "")]:
                if newLegends[flow.replace("__Building" + str(building), "")].replace("_in", "") in list:
                    production[newLegends[flow.replace("__Building" + str(building), "")].replace("_in", "")][2] -= sum(elBus[flow])
                else:
                    production[newLegends[flow.replace("__Building" + str(building), "")].replace("_in", "")] = [0, 0, -sum(elBus[flow])]
                    list.append(newLegends[flow.replace("__Building" + str(building), "")].replace("_in", ""))
            elif "_out" in newLegends[flow.replace("__Building" + str(building), "")]:
                if newLegends[flow.replace("__Building" + str(building), "")].replace("_out", "") in list:
                    production[newLegends[flow.replace("__Building" + str(building), "")].replace("_out", "")][2] += sum(elBus[flow])
                else:
                    production[newLegends[flow.replace("__Building" + str(building), "")].replace("_out", "")] = [0, 0, sum(elBus[flow])]
                    list.append(newLegends[flow.replace("__Building" + str(building), "")].replace("_out", ""))
        elif "Storage" not in flow and "Demand" not in flow:
            production[newLegends[flow.replace("__Building"+str(building), "")]] = [0, 0, sum(elBus[flow])]

            # specific for negative flows
            if "HP_SH" in flow or "HP_DHW" in flow:
                alpha += sum(elBus[flow])

    production["HP"] = [0, 0, -alpha]

    for flow in shBus.keys():
        if "Storage" in flow and "out" in newLegends[flow.replace("__Building"+str(building), "")]:
            storage[newLegends[flow.replace("__Building"+str(building), "")]] = [0, sum(shBus[flow]), 0]
        elif "Storage" not in flow and "Demand" not in flow:
            production[newLegends[flow.replace("__Building"+str(building), "")]] = [0, sum(shBus[flow]), 0]

    for flow in dhwBus.keys():
        if "Storage__" in flow and "out" in newLegends[flow.replace("__Building"+str(building), "")]:
            storage[newLegends[flow.replace("__Building"+str(building), "")]] = [sum(dhwBus[flow]), 0, 0]
        elif "Storage__" not in flow and "Demand" not in flow:
            production[newLegends[flow.replace("__Building"+str(building), "")]] = [sum(dhwBus[flow]), 0, 0]

    costs = costs.transpose()
    costs = costs.rename(index={0: building})

    env = env.transpose()
    for i in env.columns:
        for j in newLegends.keys():
            if i.replace("__Building"+str(building), "") in j:
                env = env.rename(columns={i: newLegends[j]})

    production = pd.DataFrame.from_dict(production, orient='index')
    production = production.transpose()

    storage = pd.DataFrame.from_dict(storage, orient='index', columns=["dhw", "sh", "elec"])
    storage = storage.transpose()

    fig = plt.figure(figsize=(18, 8))
    gs = gridspec.GridSpec(1, 4, width_ratios=[3, 4, 8, 3], wspace=0.03)

    ax0 = plt.subplot(gs[0])
    costs_colors = [toColor(COLORS, x) for x in costs.columns]
    bp0 = costs.plot(ax=ax0, kind='barh', color=costs_colors, stacked=True, linewidth=0)

    ax1 = plt.subplot(gs[1])
    env_colors = [toColor(COLORS, x) for x in env.columns]
    bp1 = env.plot(ax=ax1, kind='barh', color=env_colors, stacked=True, linewidth=0)

    ax2 = plt.subplot(gs[2])
    production_colors = [toColor(COLORS, x) for x in production.columns]
    bp2 = production.plot(ax=ax2, kind='barh', color=production_colors, stacked=True, linewidth=0, width=.5)

    ax3 = plt.subplot(gs[3])
    storage_colors = [toColor(COLORS, x) for x in storage.columns]
    bp3 = storage.plot(ax=ax3, kind='barh', color=storage_colors, stacked=True, linewidth=0)

    # remove scenario names from other bar plots
    for ax in [ax1, ax2]:
        ax.set_yticklabels('')
    for ax in [ax2, ax3]:
        groupHbarPlots(ax, group_size=3, inner_sep=0.01)
    ax3.yaxis.tick_right()

    # set limits and ticks for both axes
    for ax in [ax0, ax1, ax2, ax3]:
        ax.yaxis.grid(False)
        ax.xaxis.grid(True, 'major', linestyle='-')
        ax.xaxis.set_ticks_position('none')
        ax.yaxis.set_ticks_position('none')

        # group 1,000,000 with commas
        xmin, xmax = ax.get_xlim()
        if xmax > 90 or xmin < -90:
            group_thousands_and_skip_first = tkr.FuncFormatter(
                lambda x, pos: '' if pos == 0 else '{:0,d}'.format(int(x)))
            ax.xaxis.set_major_formatter(group_thousands_and_skip_first)
        else:
            skip_lowest = tkr.FuncFormatter(
                lambda x, pos: '' if pos == 0 else x)
            ax.xaxis.set_major_formatter(skip_lowest)

        # legend
        # set style arguments
        legend_style = {'frameon': False,
                        'loc': 'lower center',
                        'ncol': 2,
                        'bbox_to_anchor': (0.5, .99)}
        # get handels and labels, remove duplicate labels
        handles, labels = deduplicateLegend(*ax.get_legend_handles_labels())
        # set legend to use those
        lg = ax.legend(handles, labels, **legend_style)
        # finally, remove lines from patches
        plt.setp(lg.get_patches(), linewidth=0)

    ax0.set_xlabel('Total costs (CHF)')
    ax1.set_xlabel('Total environmental impacts (kgCo2eq)')
    ax2.set_xlabel('Total energy produced (kWh)')
    ax3.set_xlabel('Retrieved energy (kWh)')
    return fig, costs, env, production, storage


def resultingDataDiagramLoop(elec, sh, dhw, costs, env, colors, buildings):
    """
    Function inspired from the URBS platform https://github.com/ojdo/urbs/blob/1house/comp.py
    Function plotting the graph comparing the different buildings/scenarios on costs, energy produced and energy
    retrieved from storages
    :param elec: list of dict type, optimization results
    :param sh: list of dict type, optimization results
    :param dhw: list of dict type, optimization results
    :param costs: list of dict type, optimization results
    :param env: list of dict type, optimization results
    :param colors: list type, different colors for the different components of the system
    :param buildings: list of str type, name of the different buildings
    :return: figure created
    """
    n_costs, n_env, n_production, n_storage = resultingDataDiagram(elec[0], sh[0], dhw[0], costs[0], env[0], colors, buildings[0])[1:]
    for i in range(1, len(elec)):
        a, b, c, d = resultingDataDiagram(elec[i], sh[i], dhw[i], costs[i], env[i], colors, buildings[i])[1:]
        n_costs = pd.concat([n_costs, a])
        n_env = pd.concat([n_env, b])
        n_production = pd.concat([n_production, c])
        n_storage = pd.concat([n_storage, d])

    fig = plt.figure(figsize=(18, 8))
    gs = gridspec.GridSpec(1, 4, width_ratios=[4, 3, 8, 3], wspace=0.03)

    ax0 = plt.subplot(gs[0])
    costs_colors = [toColor(colors, x) for x in n_costs.columns]
    bp0 = n_costs.plot(ax=ax0, kind='barh', color=costs_colors, stacked=True, linewidth=0)

    ax1 = plt.subplot(gs[1])
    env_colors = [toColor(colors, x) for x in n_env.columns]
    bp1 = n_env.plot(ax=ax1, kind='barh', color=env_colors, stacked=True, linewidth=0)

    ax2 = plt.subplot(gs[2])
    production_colors = [toColor(colors, x) for x in n_production.columns]
    bp2 = n_production.plot(ax=ax2, kind='barh', color=production_colors, stacked=True, linewidth=0, width=.5)

    ax3 = plt.subplot(gs[3])
    storage_colors = [toColor(colors, x) for x in n_storage.columns]
    bp3 = n_storage.plot(ax=ax3, kind='barh', color=storage_colors, stacked=True, linewidth=0)

    # remove scenario names from other bar plots
    for ax in [ax1, ax2, ax3]:
        ax.set_yticklabels('')
    for ax in [ax2, ax3]:
        groupHbarPlots(ax, group_size=3, inner_sep=0.01)

    # set limits and ticks for both axes
    for ax in [ax0, ax1, ax2, ax3]:
        ax.yaxis.grid(False)
        ax.xaxis.grid(True, 'major', linestyle='-')
        ax.xaxis.set_ticks_position('none')
        ax.yaxis.set_ticks_position('none')

        # group 1,000,000 with commas
        xmin, xmax = ax.get_xlim()
        if xmax > 90 or xmin < -90:
            group_thousands_and_skip_first = tkr.FuncFormatter(
                lambda x, pos: '' if pos == 0 else '{:0,d}'.format(int(x)))
            ax.xaxis.set_major_formatter(group_thousands_and_skip_first)
        else:
            skip_lowest = tkr.FuncFormatter(
                lambda x, pos: '' if pos == 0 else x)
            ax.xaxis.set_major_formatter(skip_lowest)

        # legend
        # set style arguments
        legend_style = {'frameon': False,
                        'loc': 'lower center',
                        'ncol': 2,
                        'bbox_to_anchor': (0.5, .99)}
        # get handels and labels, remove duplicate labels
        handles, labels = deduplicateLegend(*ax.get_legend_handles_labels())
        # set legend to use those
        lg = ax.legend(handles, labels, **legend_style)
        # finally, remove lines from patches
        plt.setp(lg.get_patches(), linewidth=0)

    ax0.set_xlabel('Total costs (CHF)')
    ax1.set_xlabel('Total environmental impacts (kgCo2eq)')
    ax2.set_xlabel('Total energy produced (kWh)\n First line: electricity, Second line: space heating,\n Third line: domestic hot water')
    ax3.set_xlabel('Retrieved energy (kWh)\n First line: electricity, Second line: space heating,\n Third line: domestic hot water')

    return fig

def resultingDataDemandDiagram(elBus, shBus, dhwBus, COLORS, building, newLegends):
    """
    Function inspired from the URBS platform https://github.com/ojdo/urbs/blob/1house/comp.py
    Function plotting the different results of the optimization. First, costs will be plotted, then the energy produced,
    comparing energy for electricity bus, sh and dhw bus, and finally the retrieved energy from the storages.
    :param elBus: dict type, results from the optimization applied to one bus. Called like "solph.views.node(results, bus)"
    :param shBus: dict type, results from the optimization applied to one bus. Called like "solph.views.node(results, bus)"
    :param dhwBus: dict type, results from the optimization applied to one bus. Called like "solph.views.node(results, bus)"
    :param COLORS: list type, different colors for the different components of the system
    :param building: str type, name of the building
    :return: Three bar plots and the elec, sh and dhw dict created
    """
    elec = {}
    sh = {}
    dhw = {}
    if optMode == "group":
        list = []
    for flow in elBus.keys():
        if "Demand" in flow:
            elec[newLegends[flow.replace("__Building"+str(building), "")]] = [0, 0, sum(elBus[flow])]
        elif "producedElectricity" in flow or "gridElectricity" in flow:
            elec[newLegends[flow.replace("__Building" + str(building), "")]] = [0, sum(elBus[flow]), 0]
        elif optMode == "group" and "electricityLink" in flow and "_in" not in newLegends[flow.replace("__Building" + str(building), "")]:
            if newLegends[flow.replace("__Building" + str(building), "")] in list:
                elec[newLegends[flow.replace("__Building" + str(building), "")]][0] += sum(elBus[flow])
            else:
                elec[newLegends[flow.replace("__Building" + str(building), "")]] = [sum(elBus[flow]), 0, 0]
                list.append(newLegends[flow.replace("__Building" + str(building), "")])
        else:
            elec[newLegends[flow.replace("__Building"+str(building), "")]] = [sum(elBus[flow]), 0, 0]

        if "Storage" in flow and "in" in newLegends[flow.replace("__Building"+str(building), "")]:
            del elec[newLegends[flow.replace("__Building"+str(building), "")]]
        elif "electricity" not in flow[5:] and "Demand" not in flow:
            del elec[newLegends[flow.replace("__Building" + str(building), "")]]
        elif "Resource" in flow or "excess" in flow:
            del elec[newLegends[flow.replace("__Building" + str(building), "")]]

    for flow in shBus.keys():
        if "Demand" in flow:
            sh[newLegends[flow.replace("__Building" + str(building), "")]] = [0, 0, sum(shBus[flow])]
        else:
            sh[newLegends[flow.replace("__Building" + str(building), "")]] = [0, sum(shBus[flow]), 0]
        if "Storage" in flow and "in" in newLegends[flow.replace("__Building"+str(building), "")]:
            del sh[newLegends[flow.replace("__Building"+str(building), "")]]

    for flow in dhwBus.keys():
        if "Demand" in flow:
            dhw[newLegends[flow.replace("__Building" + str(building), "")]] = [0, 0, sum(dhwBus[flow])]
        else:
            dhw[newLegends[flow.replace("__Building" + str(building), "")]] = [0, sum(dhwBus[flow]), 0]

        if "Storage_" in flow:
            del dhw[newLegends[flow.replace("__Building"+str(building), "")]]

    elec = pd.DataFrame.from_dict(elec, orient='index')
    elec = elec.transpose()

    sh = pd.DataFrame.from_dict(sh, orient='index')
    sh = sh.transpose()

    dhw = pd.DataFrame.from_dict(dhw, orient='index')
    dhw = dhw.transpose()

    fig = plt.figure(figsize=(18, 8))
    gs = gridspec.GridSpec(1, 3, wspace=0.1)

    ax0 = plt.subplot(gs[0])
    elec_colors = [toColor(COLORS, x) for x in elec.columns]
    bp0 = elec.plot(ax=ax0, kind='barh', color=elec_colors, stacked=True, linewidth=0, width=.5)

    ax1 = plt.subplot(gs[1])
    sh_colors = [toColor(COLORS, x) for x in sh.columns]
    bp1 = sh.plot(ax=ax1, kind='barh', color=sh_colors, stacked=True, linewidth=0)

    ax2 = plt.subplot(gs[2])
    dhw_colors = [toColor(COLORS, x) for x in dhw.columns]
    bp2 = dhw.plot(ax=ax2, kind='barh', color=dhw_colors, stacked=True, linewidth=0)

    # remove scenario names from other bar plots
    ax0.set_yticklabels(('', '', 'Building '+str(building)))
    for ax in [ax1, ax2]:
        ax.set_yticklabels('')
    for ax in [ax0, ax1, ax2]:
        groupHbarPlots(ax, group_size=3, inner_sep=0.0)

    # set limits and ticks for both axes
    for ax in [ax0, ax1, ax2]:
        ax.yaxis.grid(False)
        ax.xaxis.grid(True, 'major', linestyle='-')
        ax.xaxis.set_ticks_position('none')
        ax.yaxis.set_ticks_position('none')

        # group 1,000,000 with commas
        xmin, xmax = ax.get_xlim()
        if xmax > 90 or xmin < -90:
            group_thousands_and_skip_first = tkr.FuncFormatter(
                lambda x, pos: '' if pos == 0 else '{:0,d}'.format(int(x)))
            ax.xaxis.set_major_formatter(group_thousands_and_skip_first)
        else:
            skip_lowest = tkr.FuncFormatter(
                lambda x, pos: '' if pos == 0 else x)
            ax.xaxis.set_major_formatter(skip_lowest)

        # legend
        # set style arguments
        legend_style = {'frameon': False,
                        'loc': 'lower center',
                        'ncol': 3,
                        'bbox_to_anchor': (0.5, .99)}
        # get handels and labels, remove duplicate labels
        handles, labels = deduplicateLegend(*ax.get_legend_handles_labels())
        # set legend to use those
        lg = ax.legend(handles, labels, **legend_style)
        # finally, remove lines from patches
        plt.setp(lg.get_patches(), linewidth=0)

    ax0.set_xlabel('Demand and energy produced\n for electricity (kWh)')
    ax1.set_xlabel('Demand and energy produced\n for space heating (kWh)')
    ax2.set_xlabel('Demand and energy produced\n for domestic hot water (kWh)')

    return fig, elec, sh, dhw

def resultingDataDemandDiagramLoop(elec, sh, dhw, colors, buildings):
    """
    Function inspired from the URBS platform https://github.com/ojdo/urbs/blob/1house/comp.py
    Function plotting the graph comparing the different buildings/scenarios on costs, energy produced and energy
    retrieved from storages
    :param elec: list of dict type, optimization results
    :param sh: list of dict type, optimization results
    :param dhw: list of dict type, optimization results
    :param colors: list type, different colors for the different components of the system
    :param buildings: list of str type, name of the different buildings
    :return: figure created
    """
    n_elec, n_sh, n_dhw = resultingDataDemandDiagram(elec[0], sh[0], dhw[0], colors, int(buildings[0]))[1:]
    for i in range(1, len(elec)):
        a, b, c = resultingDataDemandDiagram(elec[i], sh[i], dhw[i], colors, buildings[i])[1:]
        n_elec = pd.concat([n_elec, a])
        n_sh = pd.concat([n_sh, b])
        n_dhw = pd.concat([n_dhw, c])

    fig = plt.figure(figsize=(18, 8))
    gs = gridspec.GridSpec(1, 3, wspace=0.1)

    ax0 = plt.subplot(gs[0])
    elec_colors = [toColor(colors, x) for x in n_elec.columns]
    bp0 = n_elec.plot(ax=ax0, kind='barh', color=elec_colors, stacked=True, linewidth=0, width=.5)

    ax1 = plt.subplot(gs[1])
    sh_colors = [toColor(colors, x) for x in n_sh.columns]
    bp1 = n_sh.plot(ax=ax1, kind='barh', color=sh_colors, stacked=True, linewidth=0)

    ax2 = plt.subplot(gs[2])
    dhw_colors = [toColor(colors, x) for x in n_dhw.columns]
    bp2 = n_dhw.plot(ax=ax2, kind='barh', color=dhw_colors, stacked=True, linewidth=0)

    # remove scenario names from other bar plots
    liste = []
    for i in buildings:
        liste.append('')
        liste.append('')
        liste.append('Building '+str(i))
    ax0.set_yticklabels(liste)
    for ax in [ax1, ax2]:
        ax.set_yticklabels('')
    for ax in [ax0, ax1, ax2]:
        groupHbarPlots(ax, group_size=3, inner_sep=0.0)

    # set limits and ticks for both axes
    for ax in [ax0, ax1, ax2]:
        ax.yaxis.grid(False)
        ax.xaxis.grid(True, 'major', linestyle='-')
        ax.xaxis.set_ticks_position('none')
        ax.yaxis.set_ticks_position('none')

        # group 1,000,000 with commas
        xmin, xmax = ax.get_xlim()
        if xmax > 90 or xmin < -90:
            group_thousands_and_skip_first = tkr.FuncFormatter(
                lambda x, pos: '' if pos == 0 else '{:0,d}'.format(int(x)))
            ax.xaxis.set_major_formatter(group_thousands_and_skip_first)
        else:
            skip_lowest = tkr.FuncFormatter(
                lambda x, pos: '' if pos == 0 else x)
            ax.xaxis.set_major_formatter(skip_lowest)

        # legend
        # set style arguments
        legend_style = {'frameon': False,
                        'loc': 'lower center',
                        'ncol': 3,
                        'bbox_to_anchor': (0.5, .99)}
        # get handels and labels, remove duplicate labels
        handles, labels = deduplicateLegend(*ax.get_legend_handles_labels())
        # set legend to use those
        lg = ax.legend(handles, labels, **legend_style)
        # finally, remove lines from patches
        plt.setp(lg.get_patches(), linewidth=0)

    ax0.set_xlabel('Demand and energy produced\n for electricity (kWh)')
    ax1.set_xlabel('Demand and energy produced\n for space heating (kWh)')
    ax2.set_xlabel('Demand and energy produced\n for domestic hot water (kWh)')

    return fig


def getData(filepath):
    """
    Function for the results recovery from an Excel file
    :param filepath: path to the Excel containing the results of the optimization
    :return: the different dicts created during the optimization
    """
    wb = load_workbook(filepath, read_only=True, keep_links=False)
    dict_sheet = {}
    for sheet in wb.sheetnames:
        dict_sheet[sheet] = pd.read_excel(filepath, sheet_name=sheet, index_col=0, engine='openpyxl')

    return dict_sheet

def loadPlottingData(resultFilePath):
    """
    Function for loading the data from excel file into variables
    :param resultFilePath: path to the Excel containing the results of the optimization
    :return: the different variables created after reading the excel file
    """
    buses = getData(resultFilePath)
    elec_names = []
    elec_dict = []
    sh_names = []
    sh_dict = []
    dhw_names = []
    dhw_dict = []
    costs_names = []
    costs_dict = []
    env_names = []
    env_dict = []

    buildings_dict = []
    buildings_names = []
    buildings_number = []
    beta = list(buses.keys())
    for i in beta:
        alpha = []
        alpha_a = []
        building = i.split("Building")[1]  # building number
        beta_bis = beta.copy()
        for j in beta:
            if j.endswith(building) and (
                    "electricityBus" in j or "electricityInBus" in j or "electricityProdBus" in j or "spaceHeatingBus" in j or "shDemandBus" in j or "shSourceBus" in j or "domesticHotWaterBus" in j):
                if "shDemandBus" not in j and "shSourceBus" not in j:
                    alpha.append(j)
                if ("electricityBus" in j) or ("electricityInBus" in j) or ("electricityProdBus" in j):
                    # find the index of electricity bus to merge the results of electricityBus, electricityInBus and electricityProdBus
                    index = -1
                    for ind in range(len(alpha_a)):
                        if "electricityBus" in alpha_a[ind].keys()[0] or "electricityInBus" in alpha_a[ind].keys()[0] or "electricityProdBus" in alpha_a[ind].keys()[0]:
                            index = ind
                    if index != -1:
                        alpha_a[index] = pd.concat((alpha_a[index], buses[j]), axis=1)
                    else:
                        alpha_a.append(buses[j])
                elif ("spaceHeatingBus" in j) or ("shDemandBus" in j) or ("shSourceBus" in j):
                    # find the index of spaceHeatingBus to merge the results of spaceHeatingBus, shDemandBus and shSourceBus
                    index = -1
                    for ind in range(len(alpha_a)):
                        if "spaceHeatingBus" in alpha_a[ind].keys()[0] or "shDemandBus" in alpha_a[ind].keys()[0] or "shSourceBus" in alpha_a[ind].keys()[0]:
                            index = ind
                    if index != -1:
                        alpha_a[index] = pd.concat((alpha_a[index], buses[j]), axis=1)
                        flow = f"(('spaceHeating__Building{building}', 'shDemandBus__Building{building}'), 'flow')"
                        if flow in alpha_a[index].columns:
                            alpha_a[index].pop(flow)
                        flow = f"(('shSourceBus__Building{building}', 'shSource__Building{building}'), 'flow')"
                        if flow in alpha_a[index].columns:
                            alpha_a[index].pop(flow)
                    else:
                        index = len(alpha_a)
                        alpha_a.append(buses[j])
                        flow = f"(('spaceHeating__Building{building}', 'shDemandBus__Building{building}'), 'flow')"
                        if flow in alpha_a[index].columns:
                            alpha_a[index].pop(flow)
                        flow = f"(('shSourceBus__Building{building}', 'shSource__Building{building}'), 'flow')"
                        if flow in alpha_a[index].columns:
                            alpha_a[index].pop(flow)
                else:
                    alpha_a.append(buses[j])
                beta_bis.remove(j)
        if alpha != []:
            buildings_dict.append(alpha_a)
            buildings_names.append(alpha)
            buildings_number.append(building)
        for a in alpha:
            beta.remove(a)

    for i in buses.keys():
        tt = {}
        b = int(i.split("Building")[1])  # building number
        if ("electricityBus" in i) or ("electricityInBus" in i) or ("electricityProdBus" in i):
            elec_names.append(i)
            if b <= len(
                    elec_dict):  # to merge the data of the two electricity buses of the same  (elec_dict[0] should have all the electricity related data of building 1, and so on...)
                elec_dict[b - 1] = pd.concat((elec_dict[b - 1], buses[i]), axis=1)
            else:
                elec_dict.append(buses[i])
        if ("spaceHeating" in i) or ("shDemand" in i) or ("shSource" in i):
            if "shDemand" not in i and "spaceHeating" not in i:
                sh_names.append(i)
                sh_dict.append(buses[i])
            else:
                ind = len(sh_names) - 1
                sh_dict[ind] = pd.concat((sh_dict[ind], buses[i]), axis=1)

                flow = f"(('spaceHeating__Building{b}', 'shDemandBus__Building{b}'), 'flow')"
                if flow in sh_dict[ind].columns:
                    sh_dict[ind].pop(flow)
                flow = f"(('shSourceBus__Building{b}', 'shSource__Building{b}'), 'flow')"
                if flow in sh_dict[ind].columns:
                    sh_dict[ind].pop(flow)

        if "domesticHotWater" in i:
            dhw_names.append(i)
            dhw_dict.append(buses[i])
        if "costs" in i:
            costs_names.append(i)
            costs_dict.append(buses[i])
        if "env_impacts" in i:
            env_names.append(i)
            env_dict.append(buses[i])

    return buses, elec_names, elec_dict, sh_names, sh_dict, dhw_names, dhw_dict, costs_names, costs_dict, env_names, env_dict,\
           buildings_dict, buildings_names, buildings_number

def createPlot(resultFilePath, basePath, plotLevel, plotType, flowType, plotAnnualHorizontalBar, newLegends):
    # load the plotting data from excel file into variables
    buses, elec_names, elec_dict, sh_names, sh_dict, dhw_names, dhw_dict, costs_names, costs_dict, env_names, env_dict, \
    buildings_dict, buildings_names, buildings_number = loadPlottingData(resultFilePath)

    if flowType == "all":
        if plotType == "energy balance":
            names = elec_names + sh_names + dhw_names
        elif plotType == "bokeh":
            names = buildings_names
            dict = buildings_dict
    elif flowType == "electricity":
        names = elec_names
        if plotType == "bokeh":
            dict = elec_dict
    elif flowType == "space heat":
        names = sh_names
        if plotType == "bokeh":
            dict = sh_dict
    elif flowType == "domestic hot water":
        names = dhw_names
        if plotType == "bokeh":
            dict = dhw_dict
    else:
        raise ValueError("Illegal value for the parameter flow type")

    months = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]

    if plotType == "energy balance":
        if plotLevel == "allMonths":
            for i in names:
                if "electricityBus" not in i:
                    monthlyBalance(buses[i], i, newLegends)
    elif plotType == "bokeh":
        ncols = 2
        if plotLevel in months:
            if flowType == "all":
                for i in range(len(buildings_names)):
                    tempDict = dict[i]
                    for j in range(len(tempDict)):
                        tempDict[j] = tempDict[j][tempDict[j].index.strftime('%b') == plotLevel]
                    dict[i] = tempDict
            else:
                for i in range(len(dict)):
                    dict[i] = dict[i][dict[i].index.strftime('%b') == plotLevel]

        elif any(chr.isdigit() for chr in plotLevel):
            if isoparse(plotLevel):
                if flowType == "all":
                    for i in range(len(buildings_names)):
                        tempDict = dict[i]
                        for j in range(len(tempDict)):
                            tempDict[j] = tempDict[j][tempDict[j].index.date == datetime.strptime(plotLevel, "%Y-%m-%d").date()]
                        dict[i] = tempDict
                else:
                    for i in range(len(dict)):
                        dict[i] = dict[i][dict[i].index.date == datetime.strptime(plotLevel, "%Y-%m-%d").date()]

        if flowType == "all":
            plotsHourly = []
            plotsDaily = []
            for i in range(len(buildings_names)):
                plotsH, plotsD = hourlyDailyPlot(dict[i], names[i], Category20_12, newLegends)
                plotsHourly.extend(plotsH)
                plotsDaily.extend(plotsD)
        else:
            plotsHourly, plotsDaily = hourlyDailyPlot(dict, names, Set1_9, newLegends)

        if not os.path.exists(basePath):
            os.makedirs(basePath)

        output_file(os.path.join(basePath,"HourlyBokehPlots.html"))
        grid = gridplot(plotsHourly, ncols=ncols, plot_width=850, plot_height=500, sizing_mode="fixed")
        show(grid)
        if not any(chr.isdigit() for chr in plotLevel):
            output_file(os.path.join(basePath,"DailyBokehPlots.html"))
            grid = gridplot(plotsDaily, ncols=ncols, plot_width=850, plot_height=500, sizing_mode="fixed")
            show(grid)
    else:
        raise ValueError("Illegal value for the parameter plot type")

    if plotAnnualHorizontalBar == True:
        my_colors = {
            'Battery_out': (100, 160, 200),
            'Battery_in': (100, 160, 200),
            'Storage_sh_out': (196, 7, 27),
            'Storage_dhw_out': (196, 7, 27),
            'HP_dhw': (62, 173, 0),
            'HP_sh': (62, 173, 0),
            'HP': (62, 173, 0),
            'CHP_elec': (0, 101, 189),
            'CHP_sh': (0, 101, 189),
            'CHP_dhw': (128, 0, 128),
            'CHP': (0, 101, 189),
            'Gas_sh': (0, 110, 120),
            'GasBoiler': (0, 110, 120),
            'PV_elec': (0, 128, 90),
            'SolarCollector': (188, 128, 90),
            'Inputs': (252, 93, 93),
            'Operation': (252, 93, 93),
            'Investment': (0, 119, 138),
            'Feed-in': (0, 215, 203),
            'Demand_elec': (131, 166, 151),
            'Demand_sh': (131, 166, 151),
            'Demand_dhw': (131, 166, 151),
            'Grid_purchase': (237, 127, 16)
        }
        if optMode == "group":
            my_colors['electricityLink_in'] = (255, 215, 0)
            my_colors['electricityLink_out'] = (255, 215, 0)
            my_colors['electricityLink'] = (255, 215, 0)
            
        COLORS = {}
        for name, color in my_colors.items():
            COLORS[name] = color
        """
        for i in range(len(buildings_names)):
            fig1 = resultingDataDiagram(elec_dict[i], sh_dict[i], dhw_dict[i], costs_dict[i], env_dict[i], COLORS, buildings_number[i])[0]
            fig2 = resultingDataDemandDiagram(elec_dict[i], sh_dict[i], dhw_dict[i], COLORS, buildings_number[i])[0]
        """
        fig3 = resultingDataDiagramLoop(elec_dict, sh_dict, dhw_dict, costs_dict, env_dict, COLORS, buildings_number)
        fig4 = resultingDataDemandDiagramLoop(elec_dict, sh_dict, dhw_dict, COLORS, buildings_number)
        plt.show()


def plot(excelFileName, figureFilePath, plotLevel, plotType, flowType, plotAnnualHorizontalBar=False):
    #####################################
    ########## Classic plots  ###########
    #####################################
    # New legends need to be defined by hand for each new flow added to the energy network
    newLegends = {
        "(('naturalGasResource', 'naturalGasBus'), 'flow')": "NaturalGas",
        "(('electricityBus', 'excesselectricityBus'), 'flow')": "Feed-in",
        "(('electricityProdBus', 'electricalStorage'), 'flow')": "Battery_in",
        "(('electricityInBus', 'electricityDemand'), 'flow')": "Demand_elec",
        "(('electricityInBus', 'HP'), 'flow')": "HP",
        "(('CHP', 'electricityProdBus'), 'flow')": "CHP_elec",
        "(('electricalStorage', 'electricityBus'), 'flow')": "Battery_out",
        "(('electricitySource', 'electricityBus'), 'flow')": "Elect_produced (not stored)",
        "(('electricityResource', 'gridBus'), 'flow')": "Grid_purchase",
        "(('dhwStorage', 'domesticHotWaterBus'), 'flow')": "Storage_dhw_out",
        "(('dhwStorageBus', 'dhwStorage'), 'flow')": "Storage_dhw_in",
        "(('domesticHotWaterBus', 'domesticHotWaterDemand'), 'flow')": "Demand_dhw",
        "(('HP', 'dhwStorageBus'), 'flow')": "HP_dhw",
        "(('GWHP', 'dhwStorageBus'), 'flow')": "GWHP_dhw",
        "(('CHP', 'dhwStorageBus'), 'flow')": "CHP_dhw",
        "(('shSource', 'spaceHeatingBus'), 'flow')": "SH_produced (not stored)",
        "(('shStorage', 'spaceHeatingBus'), 'flow')": "Storage_sh_out",
        "storage_content": "Storage_content",
        "(('shSourceBus', 'shStorage'), 'flow')": "Storage_sh_in",
        "(('spaceHeatingBus', 'spaceHeating'), 'flow')": "SH_direct_to_load",
        "(('shDemandBus', 'spaceHeatingDemand'), 'flow')": "Demand_sh",
        "(('CHP', 'shSourceBus'), 'flow')": "CHP_sh",
        "(('GasBoiler', 'shSourceBus'), 'flow')": "Gas_sh",
        "(('GasBoiler', 'dhwStorageBus'), 'flow')": "Gas_dhw",
        "(('HP', 'shSourceBus'), 'flow')": "HP_sh",
        "(('GWHP', 'shSourceBus'), 'flow')": "GWHP_sh",
        "(('electricityBus', 'producedElectricity'), 'flow')": "Self_consumption",
        "(('gridBus', 'gridElectricity'), 'flow')": "Electricity_grid",
        "(('pv', 'electricityProdBus'), 'flow')": "PV_elec",
        "(('solarCollector', 'dhwStorageBus'), 'flow')": "SolarCollector",
        "(('electricityInBus', 'solarCollector'), 'flow')": "SolarCollector",
        "(('gridElectricity', 'electricityInBus'), 'flow')": "Electricity_grid",
        "(('producedElectricity', 'electricityInBus'), 'flow')": "Self_consumption",
        "(('electricityProdBus', 'electricitySource'), 'flow')": "Battery_bypass"
    }
    newLegends["(('electricityBus', 'electricityLink'), 'flow')"] = "electricityLink_in"
    newLegends["(('electricityLink', 'electricityInBus'), 'flow')"] = "electricityLink_out"
    newLegends["(('spaceHeatingBus', 'shLink'), 'flow')"] = "shLink_in"
    newLegends["(('shLink', 'shDemandBus'), 'flow')"] = "shLink_out"

    createPlot(excelFileName, figureFilePath, plotLevel, plotType, flowType, plotAnnualHorizontalBar, newLegends)


if __name__ == '__main__':

    optMode = "group"  # parameter defining whether the results file corresponds to "indiv" or "group" optimization
    numberOfBuildings = 4
    plotOptim = 1  # defines the number of the optimization to plot
    plotLevel = "allMonths"  # permissible values (for energy balance plot): "allMonths" {for all months}
    # or specific month {"Jan", "Feb", "Mar", etc. three letter abbreviation of the month name}
    # or specific date {format: YYYY-MM-DD}
    plotType = "energy balance"  # permissible values: "energy balance", "bokeh"
    flowType = "electricity"  # permissible values: "all", "electricity", "space heat", "domestic hot water"
    plotAnnualHorizontalBar = False  # determines whether the annual horizontal bar is plot or not

    plot(os.path.join(r"..\data\Results", f"results{numberOfBuildings}_{plotOptim}_{optMode}.xlsx"), r"..\data\Figures", plotLevel, plotType, flowType, plotAnnualHorizontalBar)
